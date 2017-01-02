import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,  Font



UNIT = 'F'
s1mme_aggr_query = []
diameter_aggr_query = []
gngi_aggr_query = []
s1mme_aggr_query_times = []
diameter_aggr_query_times = []
gngi_aggr_query_times = []
s1mme_aggr_query_rate = []
diameter_aggr_query_rate = []
gngi_aggr_query_rate = []


wb = openpyxl.load_workbook(filename='./LTE_KPIs_up.xlsx')
# grab the  worksheet
ws1 = wb['KPIs']
resulting_wb = Workbook()
ws = resulting_wb.create_sheet('counts')
ws = resulting_wb.create_sheet('times')
ws = resulting_wb.create_sheet('rate')
ws_count = resulting_wb.get_sheet_by_name('counts')
ws_times = resulting_wb.get_sheet_by_name('times')
ws_rate = resulting_wb.get_sheet_by_name('rate')

date1 = str(ws1['E3'].value)[:11]
date2 = str(ws1['F3'].value)[:11]


def create_counts_select(sheet, index):
    if is_transaction(sheet, sheet['G' + str(index)]):
        success = 'AND ' + TRANSACTION_CAUSE_TYPE + ' is NULL'
    elif is_transaction(sheet, sheet['G' + str(index)]) == False:
        success = ' AND ' + TRANSACTION_CAUSE_TYPE + ' is not NULL'
    else:
        success = ''
    protocol = protocol_id(sheet, sheet['H' + str(index)])
    i = 0
    if protocol == '-1':
        while protocol == '-1':
            i += 1
            protocol = protocol_id(sheet, sheet['H' + str(index - i)])
            if index - 1 == 0:
                break
            else:
                continue
    else:
        protocol = protocol

    query = 'When {0} = {1} AND {2} = {3} {4} THEN "{5}"'\
    .format(PROTOCOL_WHEN, protocol,
    TRANSACTION_TYPE, transaction_type_ID(sheet, sheet['H' + str(index)]), success,
    kpi_name(sheet['D' + str(index)]))
    return (str(query))


# method to get KPI name value
def kpi_name(cell):
    return cell.value
# method which determine whether transaction type successful or not
def is_transaction(sheet, cell):
    raw_data = cell.value
    pattern1 = re.compile(r'[^A-Z][^un]Successesful')
    pattern2 = re.compile(r'[^A-Z]unSuccessesful')
    success = pattern1.findall(str(raw_data))
    unsuccess = pattern2.findall(str(raw_data))
    if success:
        return True
    elif unsuccess:
        return False
    else:
        return
# testing method
#print('success or not - {0}'.format(str(is_transaction(ws1, ws1['G9']))))

def getValueWithMergeLookup(sheet, cell):
    idx = cell.coordinate
    for range_ in sheet.merged_cell_ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(range_))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return sheet.cell(merged_cells[0][0]).value

    return sheet.cell(idx).value


def define_protocol_name(sheet, cell):
    cell_value = getValueWithMergeLookup(sheet, cell)
    global AGGREGATE_TABLE
    global PROTOCOL_WHEN
    global TRANSACTION_TYPE
    global TRANSACTION_CAUSE_TYPE
    global S1MNE_AGGR
    global DIAMETER_AGGR
    global GNGI_AGGR
    if cell_value in ('S1AP', 'EPS NAS', 'SGsAP'):
        AGGREGATE_TABLE = 's1mme_aggr'
        PROTOCOL_WHEN = 'S1AP_SGS_PROTOCOL_ID'
        TRANSACTION_TYPE = 'S1AP_SGS_TRANS_TYPE'
        TRANSACTION_CAUSE_TYPE = 'S1AP_SGS_TRANS_CAUSE_TYPE'
        S1MNE_AGGR = True
    elif cell_value == 'DIAMETER':
        AGGREGATE_TABLE = 'diameter_aggr'
        PROTOCOL_WHEN = 'PROTOCOLID'
        TRANSACTION_TYPE = 'TRANSACTIONTYPE'
        TRANSACTION_CAUSE_TYPE = 'CAUSETYPE'
        DIAMETER_AGGR = True
    elif cell_value == 'GTPv2':
        AGGREGATE_TABLE = 'gngi_aggr'
        PROTOCOL_WHEN = 'PROTOCOLID'
        TRANSACTION_TYPE = 'TRANS_STATS_TYPE'
        TRANSACTION_CAUSE_TYPE = 'CAUSE_TYPE'
        GNGI_AGGR = True
    else:
        cell_value = '0'
    return cell_value


def protocol_id(sheet, cell):
    raw_data = getValueWithMergeLookup(sheet, cell)
    pattern1 = re.compile(r'[^a-z]protocol[\s]ID[\s][=][\s]\d{1,}')
    num = pattern1.findall(str(raw_data))
    if num:
        return re.sub('\D', '', str(num))
    else:
        return '-1'


def transaction_type_ID(sheet, cell):
    raw_data = getValueWithMergeLookup(sheet, cell)
    pattern2 = re.compile(r'Transaction[\s]Type[\s]ID[\s][=][\s]\d{1,}')
    num = pattern2.findall(str(raw_data))
    if num:
        return re.sub('\D', '', str(num))
    else:
        return '-1'

# print stored queries to the Excel sheet
def write_query_to_file(pointer, list, table, sheet, end_statement):
    # apply style
    ca = sheet['A' + str(pointer + 1)]
    ct = sheet['A' + str(pointer)]
    al = Alignment(vertical='center', text_rotation=90)
    txt = Alignment(vertical='center', wrap_text=True)
    ca.alignment = al
    ct.alignment = txt
    sheet['A' + str(pointer)] = 'aggregate table'
    # merge for table name
    sheet.merge_cells(start_row=pointer+1, start_column=1, end_row=len(list)+pointer, end_column=1)
    sheet['A' + str(pointer +1)] = str(table)
    sheet['B' + str(pointer + 1)] = 'select case'
    sheet['C' + str(len(list) + pointer + 1)] = end_statement
    # only for count 'Sum(cnt) string should be printed'
    if sheet.title == 'counts':
        sheet['C' + str(len(list) + pointer + 2)] = 'Sum(cnt) cnt'
        sheet['C' + str(len(list) + pointer + 3)] = 'From {0}'.format(table)
        sheet['C' + str(len(list) + pointer + 4)] = 'Where date between {0} and {1}'.format(date1, date2)
    else:
        sheet['C' + str(len(list) + pointer + 2)] = 'From {0}'.format(table)
        sheet['C' + str(len(list) + pointer + 3)] = 'Where date between {0} and {1}'.format(date1, date2)

    for query in list:
        pointer += 1
        sheet.merge_cells(start_row=pointer, start_column=3, end_row=pointer, end_column=7)
        sheet['C' + str(pointer)] = str(query).split('THEN')[0]
        sheet['H' + str(pointer)] = 'THEN'
        sheet['I' + str(pointer)] = str(query).split('THEN')[1] + ','

    # overwrite cell to get rid of last comma
    sheet['I' + str(pointer)] = str(query).split('THEN')[1]
    return


# global loop Unit column to define Select type and store queries in the Lists
for i in range(9, 183):
    index = UNIT+ (str(i))
    type_of_select = ws1[str(index)].value
    GNGI_AGGR, DIAMETER_AGGR, S1MNE_AGGR = False, False, False
    define_protocol_name(ws1, ws1['A' + str(i)])
    if type_of_select == '#':
        if S1MNE_AGGR:
            s1mme_aggr_query.append(str(create_counts_select(ws1, i)))
        elif DIAMETER_AGGR:
            diameter_aggr_query.append(str(create_counts_select(ws1, i)))
        elif GNGI_AGGR:
            gngi_aggr_query.append(str(create_counts_select(ws1, i)))
        else:
            continue
    elif type_of_select == '%':
        if S1MNE_AGGR:
            s1mme_aggr_query_rate.append(str(create_counts_select(ws1, i)))
        elif DIAMETER_AGGR:
            diameter_aggr_query_rate.append(str(create_counts_select(ws1, i)))
        elif GNGI_AGGR:
            gngi_aggr_query_rate.append(str(create_counts_select(ws1, i)))
        else:
            continue
    elif type_of_select == 'miliseconds':
        if S1MNE_AGGR:
            s1mme_aggr_query_times.append(str(create_counts_select(ws1, i)))
        elif DIAMETER_AGGR:
            diameter_aggr_query_times.append(str(create_counts_select(ws1, i)))
        elif GNGI_AGGR:
            gngi_aggr_query_times.append(str(create_counts_select(ws1, i)))
        else:
            continue
    else:
        continue


# call function to write count to Excel
ending = 'End kpi_name,'
write_query_to_file(1, s1mme_aggr_query, 's1mne_aggr', ws_count, ending)
write_query_to_file(int(len(s1mme_aggr_query)+6), diameter_aggr_query, 'diameter_aggr', ws_count,ending)
write_query_to_file(int(len(diameter_aggr_query)+ len(s1mme_aggr_query)+12), gngi_aggr_query, 'gngi_aggr', ws_count, ending)

# call function to write times to Excel
ending = 'end as kpi_name, 1000*sum(time)/sum(cnt)'
write_query_to_file(1, s1mme_aggr_query_times, 's1mne_aggr', ws_times, ending)
write_query_to_file(int(len(s1mme_aggr_query_times)+6), diameter_aggr_query_times, 'diameter_aggr', ws_times, ending)
write_query_to_file(int(len(diameter_aggr_query_times)+ len(s1mme_aggr_query_times)+12), gngi_aggr_query_times, 'gngi_aggr', ws_times, ending)

# call function to write rate to Excel
ending = 'max(case S1AP_SGS_TRANS_CAUSE_TYPE when  is not null then cnt end )' \
         '\n / nullif(max(case when S1AP_SGS_TRANS_CAUSE_TYPE is null then cnt end),0) as result'
write_query_to_file(1, s1mme_aggr_query_rate, 's1mne_aggr', ws_rate, ending)
ending = 'max(case CAUSETYPE when  is not null then cnt end )' \
         '\n / nullif(max(case when CAUSETYPE is null then cnt end),0) as result'
write_query_to_file(int(len(s1mme_aggr_query_rate)+6), diameter_aggr_query_rate, 'diameter_aggr', ws_rate, ending)
ending = 'max(case TRANS_CAUSE_TYPE when  is not null then cnt end )' \
         '\n / nullif(max(case when TRANS_CAUSE_TYPE is null then cnt end),0) as result'
write_query_to_file(int(len(diameter_aggr_query_rate)+ len(s1mme_aggr_query_rate)+12), gngi_aggr_query_rate, 'gngi_aggr', ws_rate, ending)


resulting_wb.save('outcome.xlsx')

# # Save the file
#wb.save("sample.xlsx")